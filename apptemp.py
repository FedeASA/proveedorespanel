"""
app.py  —  Panel RMA Proveedores
Fase 3 completa: Tabla 1 (Por Enviar) + Tabla 2 (Enviados) + Tabla 3 (Finalizados).

Optimizaciones de rendimiento:
  • Tabla 1: batch query Drive (1 llamada API para todas las carpetas de proveedor).
             Solo lee archivos directos en la carpeta (no /procesado, no /finalizado).
             Conteo de artículos cacheado por (file_id + modifiedTime): no re-descarga.
  • Tabla 2/3: leen de Google Sheets (muy rápido). Metadatos de archivo cacheados.
  • Refresh incremental: al presionar Refresh busca solo archivos modificados
             desde el último escaneo completo y actualiza esas entradas.
"""

import re
import io
from datetime import date, datetime

import pandas as pd
import streamlit as st

from lib.drive_ops import (
    asegurar_subcarpeta,
    contar_articulos_excel_cached,
    extraer_file_id_de_link,
    extraer_folder_id_de_link,
    leer_excel_completo,
    listar_carpetas_invalidas,
    listar_carpetas_proveedor,
    listar_excels_todas_carpetas,
    listar_excels_modificados_desde,
    mover_archivo,
    obtener_nombre_archivo,
    obtener_parent_desde_drive,
    renombrar_archivo,
    validar_nombre_excel,
)
from lib.google_client import get_google_clients
from lib.sheets_ops import (
    actualizar_registro,
    clear_cache,
    leer_proveedores,
    leer_registros,
    upsert_registro,
)
from lib.airtable_ops import (
    listar_proveedores_unicos,
    obtener_registros_proveedor,
    marcar_como_capturados,
)

# ══════════════════════════════════════════════════════════
#  CONFIGURACIÓN
# ══════════════════════════════════════════════════════════

st.set_page_config(page_title="Panel RMA Proveedores", layout="wide")

st.markdown(
    """
    <style>
        .block-container {
            max-width: 100% !important;
            padding-left: 2rem !important;
            padding-right: 2rem !important;
            padding-top: 4rem;
        }
        div[data-testid="stExpander"] { border: 1px solid #444; margin-bottom: 1rem; }
        .stDataTable td, .stDataTable th {
            font-family: sans-serif !important;
            font-size: 14px !important;
            font-weight: 400 !important;
            border-right: 1px solid #444 !important;
        }
        .tabla1-head-cell {
            font-weight: 700;
            color: #e8e8e8;
            padding: 0.2rem 0 0.35rem 0;
        }
        .tabla1-header-divider { border-bottom: 1px solid #444; margin-bottom: 0.2rem; }
        .tabla1-row-divider    { border-bottom: 1px solid #333; margin: 0.2rem 0 0.35rem 0; }
        .badge-nc   { background:#28a745; color:white; padding:2px 8px; border-radius:4px; font-size:12px; }
        .badge-cambio { background:#17a2b8; color:white; padding:2px 8px; border-radius:4px; font-size:12px; }
        .badge-rechazado { background:#dc3545; color:white; padding:2px 8px; border-radius:4px; font-size:12px; }
    </style>
    """,
    unsafe_allow_html=True,
)

SHEET_ID = "1wBkIvtk_KDcPB3Jt1vBcWdVfpRPvzAevX1ooBQ550sI"
DRIVE_ROOT_FOLDER_ID = "1ypABIZBCkvsRED1I8tYE5oxAOT32QTh6"

# ══════════════════════════════════════════════════════════
#  AUTENTICACIÓN
# ══════════════════════════════════════════════════════════

if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
    st.session_state.usuario = ""
    st.session_state.rol = ""


def login() -> None:
    st.markdown(
        "<h2 style='text-align: center;'>Control de Acceso — Panel RMA Proveedores</h2>",
        unsafe_allow_html=True,
    )
    with st.form("formulario_login"):
        usuario = st.text_input("Usuario:").strip()
        clave = st.text_input("Contraseña:", type="password").strip()
        bot = st.form_submit_button("Iniciar Sesión", use_container_width=True)
        if bot:
            try:
                us = st.secrets["USUARIOS"]
                if usuario in us and us[usuario] == clave:
                    st.session_state.autenticado = True
                    st.session_state.usuario = usuario
                    st.session_state.rol = "admin" if usuario == "admin" else "user"
                    st.rerun()
                else:
                    st.error("Usuario o contraseña incorrectos.")
            except Exception:
                st.error("Error de configuración: falta [USUARIOS] en secrets.")


if not st.session_state.autenticado:
    login()
    st.stop()

st.sidebar.write(
    f"Conectado como: **{st.session_state.usuario}** ({st.session_state.rol.upper()})"
)
if st.sidebar.button("Cerrar Sesión", use_container_width=True):
    for k in ("autenticado", "usuario", "rol"):
        st.session_state[k] = False if k == "autenticado" else ""
    st.rerun()

# ══════════════════════════════════════════════════════════
#  HELPERS DE FORMATO
# ══════════════════════════════════════════════════════════

def _to_ddmmyyyy(value: str) -> str:
    return datetime.strptime(value, "%Y-%m-%d").strftime("%d/%m/%Y")


def _to_ddmmyyyy_safe(value: str) -> str:
    if not value or str(value).strip() in ("", "None", "nan"):
        return ""
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def _slugify(value: str) -> str:
    base = value.rsplit(".", 1)[0].strip().lower()
    base = re.sub(r"\s+", "_", base)
    base = re.sub(r"[^a-z0-9_]+", "", base)
    return base or "archivo"


def _nombre_finalizado(nombre_actual: str, resolucion: str, numero_nc: str) -> str:
    base, dot, ext = nombre_actual.rpartition(".")
    if not dot:
        base, ext = nombre_actual, ""
    sufijo = {
        "NOTA DE CREDITO": f" - finalizado - NOTA DE CREDITO {numero_nc.strip()}",
        "CAMBIO": " - finalizado - CAMBIO",
        "RECHAZADO": " - finalizado - RECHAZADO",
    }.get(str(resolucion).strip(), " - finalizado")
    base_nuevo = re.sub(r"\s-\senviado\s*$", "", base, flags=re.IGNORECASE).rstrip()
    return f"{base_nuevo}{sufijo}.{ext}" if ext else f"{base_nuevo}{sufijo}"


def _badge_resolucion(res: str) -> str:
    css = {"NOTA DE CREDITO": "badge-nc", "CAMBIO": "badge-cambio", "RECHAZADO": "badge-rechazado"}
    cls = css.get(res.strip().upper(), "")
    return f'<span class="{cls}">{res}</span>' if cls else res


# ══════════════════════════════════════════════════════════
#  SCAN TABLA 1  (con cache y refresh incremental)
# ══════════════════════════════════════════════════════════

# Guardamos el resultado del último escaneo completo en session_state
# para que el refresh incremental pueda comparar y solo re-descargar
# los archivos que realmente cambiaron.

def _inicializar_scan_state():
    if "_t1_filas" not in st.session_state:
        st.session_state._t1_filas = None
        st.session_state._t1_carpetas_validas = []
        st.session_state._t1_carpetas_invalidas = []
        st.session_state._t1_excels_invalidos = []
        st.session_state._t1_last_scan_iso = None   # RFC3339 del último scan completo


_inicializar_scan_state()


@st.cache_data(ttl=120, show_spinner=False)
def _scan_tabla_1_completo():
    """
    Escaneo completo de Tabla 1.
    Cache de 120 s entre recargas de página (sin click en Refresh).
    """
    drive_client, _ = get_google_clients()
    proveedores = leer_proveedores(SHEET_ID)
    carpetas_validas = listar_carpetas_proveedor(drive_client, DRIVE_ROOT_FOLDER_ID, proveedores)
    carpetas_invalidas = listar_carpetas_invalidas(drive_client, DRIVE_ROOT_FOLDER_ID, proveedores)

    filas, excels_invalidos = _procesar_excels(drive_client, carpetas_validas)
    scan_time = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
    return filas, carpetas_validas, carpetas_invalidas, excels_invalidos, scan_time


def _procesar_excels(drive_client, carpetas_validas: list) -> tuple[list, list]:
    """
    Usa la BATCH QUERY para traer todos los Excels de todas las carpetas
    con UNA sola llamada a la API Drive.
    """
    excels_por_carpeta = listar_excels_todas_carpetas(drive_client, carpetas_validas)
    filas = []
    excels_invalidos = []
    carpetas_map = {c["folder_id"]: c for c in carpetas_validas}

    for folder_id, excels in excels_por_carpeta.items():
        carpeta = carpetas_map[folder_id]
        for excel in excels:
            validacion = validar_nombre_excel(excel["name"])
            if not validacion["valido"]:
                excels_invalidos.append(
                    {
                        "carpeta_proveedor": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                        "archivo": excel["name"],
                        "motivo": validacion["motivo_error"],
                    }
                )
                continue
            if str(validacion["proveedor"]).strip() != carpeta["nombre"].strip():
                excels_invalidos.append(
                    {
                        "carpeta_proveedor": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                        "archivo": excel["name"],
                        "motivo": "Proveedor en nombre de archivo no coincide con carpeta.",
                    }
                )
                continue

            # Conteo cacheado por (file_id + modifiedTime)
            cantidad = contar_articulos_excel_cached(
                drive_client, excel["id"], excel.get("modifiedTime", "")
            )
            filas.append(
                {
                    "id_proveedor": carpeta["id_proveedor"],
                    "proveedor_nombre": carpeta["nombre"],
                    "proveedor_mostrado": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                    "fecha_iso": validacion["fecha"],
                    "fecha": _to_ddmmyyyy(validacion["fecha"]) if validacion["fecha"] else "",
                    "cantidad_articulos": cantidad if cantidad is not None else "",
                    "archivo_nombre": excel["name"],
                    "archivo_link": excel.get("webViewLink", ""),
                    "archivo_id": excel["id"],
                    "archivo_modified": excel.get("modifiedTime", ""),
                    "folder_id": carpeta["folder_id"],
                    "folder_link": carpeta.get("folder_link", ""),
                }
            )

    filas.sort(key=lambda x: (x["id_proveedor"], x["fecha_iso"], x["archivo_nombre"]))
    return filas, excels_invalidos


def _scan_tabla_1_incremental(carpetas_validas: list, filas_actuales: list, last_scan_iso: str):
    """
    Refresh incremental: solo re-procesa archivos modificados desde last_scan_iso.
    Archivos sin cambios mantienen sus filas existentes (conteo ya cacheado).
    """
    drive_client, _ = get_google_clients()
    folder_ids = [c["folder_id"] for c in carpetas_validas]
    carpetas_map = {c["folder_id"]: c for c in carpetas_validas}

    modificados_por_carpeta = listar_excels_modificados_desde(
        drive_client, folder_ids, last_scan_iso
    )

    ids_modificados = {
        f["id"]
        for archivos in modificados_por_carpeta.values()
        for f in archivos
    }

    if not ids_modificados:
        return filas_actuales, []

    # Mantener filas de archivos NO modificados
    filas_sin_cambio = [f for f in filas_actuales if f["archivo_id"] not in ids_modificados]

    # Procesar solo los modificados
    carpetas_con_cambios = [
        c for c in carpetas_validas if c["folder_id"] in modificados_por_carpeta
        and modificados_por_carpeta[c["folder_id"]]
    ]
    # Temporalmente sobreescribimos el dict para _procesar_excels
    _tmp_by_folder = {c["folder_id"]: modificados_por_carpeta[c["folder_id"]]
                      for c in carpetas_con_cambios}
    # Reusar _procesar_excels directamente sobre el subconjunto
    filas_nuevas, excels_invalidos_nuevos = _procesar_excels_desde_dict(
        drive_client, _tmp_by_folder, carpetas_map
    )

    todas = filas_sin_cambio + filas_nuevas
    todas.sort(key=lambda x: (x["id_proveedor"], x["fecha_iso"], x["archivo_nombre"]))
    return todas, excels_invalidos_nuevos


def _procesar_excels_desde_dict(drive_client, excels_por_carpeta: dict, carpetas_map: dict):
    filas, excels_invalidos = [], []
    for folder_id, excels in excels_por_carpeta.items():
        carpeta = carpetas_map.get(folder_id)
        if not carpeta:
            continue
        for excel in excels:
            validacion = validar_nombre_excel(excel["name"])
            if not validacion["valido"]:
                excels_invalidos.append({
                    "carpeta_proveedor": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                    "archivo": excel["name"],
                    "motivo": validacion["motivo_error"],
                })
                continue
            if str(validacion["proveedor"]).strip() != carpeta["nombre"].strip():
                excels_invalidos.append({
                    "carpeta_proveedor": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                    "archivo": excel["name"],
                    "motivo": "Proveedor en nombre no coincide con carpeta.",
                })
                continue
            cantidad = contar_articulos_excel_cached(
                drive_client, excel["id"], excel.get("modifiedTime", "")
            )
            filas.append({
                "id_proveedor": carpeta["id_proveedor"],
                "proveedor_nombre": carpeta["nombre"],
                "proveedor_mostrado": f'{carpeta["id_proveedor"]} - {carpeta["nombre"]}',
                "fecha_iso": validacion["fecha"],
                "fecha": _to_ddmmyyyy(validacion["fecha"]) if validacion["fecha"] else "",
                "cantidad_articulos": cantidad if cantidad is not None else "",
                "archivo_nombre": excel["name"],
                "archivo_link": excel.get("webViewLink", ""),
                "archivo_id": excel["id"],
                "archivo_modified": excel.get("modifiedTime", ""),
                "folder_id": carpeta["folder_id"],
                "folder_link": carpeta.get("folder_link", ""),
            })
    return filas, excels_invalidos


# ══════════════════════════════════════════════════════════
#  CACHE DE METADATOS DE ARCHIVO (para Tabla 2 y 3)
# ══════════════════════════════════════════════════════════

@st.cache_data(ttl=300, show_spinner=False)
def _get_file_meta_cached(file_id: str) -> dict:
    """Cachea nombre del archivo por file_id durante 5 minutos."""
    if not file_id:
        return {"name": ""}
    try:
        drive_client, _ = get_google_clients()
        meta = (
            drive_client.files()
            .get(fileId=file_id, fields="name,webViewLink", supportsAllDrives=True)
            .execute()
        )
        return meta
    except Exception:
        return {"name": ""}


@st.cache_data(ttl=300, show_spinner=False)
def _leer_excel_detalle_cache(file_id: str):
    drive_client, _ = get_google_clients()
    return leer_excel_completo(drive_client, file_id)


# ══════════════════════════════════════════════════════════
#  DIÁLOGOS
# ══════════════════════════════════════════════════════════

@st.dialog("Detalle de artículos")
def _dialog_detalle_articulos(nombre_archivo: str, file_id: str):
    st.markdown(f"**{nombre_archivo}**")
    df = _leer_excel_detalle_cache(file_id)
    if df is None:
        st.error("No se pudo leer el Excel desde Drive.")
        return
    st.dataframe(df, use_container_width=True, hide_index=True)


@st.dialog("Marcar como ENVIADO")
def _dialog_enviado(fila: dict):
    with st.form("form_enviado"):
        numero_envio = st.text_input("Número de envío")
        fecha_envio = st.date_input("Fecha de envío", value=date.today(), format="DD/MM/YYYY")
        link_seguimiento = st.text_input("Link de seguimiento (opcional)")
        link_comprobante = st.text_input("Link de PDF / comprobante (opcional)")
        confirmar = st.form_submit_button("CONFIRMAR", type="primary")

        if confirmar:
            if not numero_envio.strip():
                st.error("Número de envío es obligatorio.")
                return

            drive_client, _ = get_google_clients()
            nombre_original = fila["archivo_nombre"]

            # Agregar sufijo " - enviado" si no está
            if " - enviado" not in nombre_original.lower():
                idx = nombre_original.rfind(".")
                nuevo_nombre = (
                    f"{nombre_original[:idx]} - enviado{nombre_original[idx:]}"
                    if idx > -1
                    else f"{nombre_original} - enviado"
                )
            else:
                nuevo_nombre = nombre_original

            renombrar_archivo(drive_client, fila["archivo_id"], nuevo_nombre)
            procesado_id = asegurar_subcarpeta(drive_client, fila["folder_id"], "procesado")
            mover_archivo(drive_client, fila["archivo_id"], procesado_id)

            file_updated = (
                drive_client.files()
                .get(fileId=fila["archivo_id"], fields="webViewLink", supportsAllDrives=True)
                .execute()
            )

            fecha_yyyymmdd = fila["fecha_iso"].replace("-", "")
            slug = _slugify(nombre_original)
            id_registro = f'{fila["id_proveedor"]}_{fecha_yyyymmdd}_{slug}'

            upsert_registro(
                id_registro,
                {
                    "Proveedor": fila["proveedor_mostrado"],
                    "Fecha_carga": fila["fecha_iso"],
                    "Cantidad_Articulos": fila["cantidad_articulos"],
                    "Link_Excel_Detalle": file_updated.get("webViewLink", fila["archivo_link"]),
                    "Estado": "ENVIADO",
                    "Numero_Envio": numero_envio.strip(),
                    "Fecha_Envio": fecha_envio.strftime("%Y-%m-%d"),
                    "Link_Seguimiento": link_seguimiento.strip(),
                    "Numero_Caso_RMA": "",
                    "Fecha_Recepcion": "",
                    "Resolucion": "",
                    "Detalle_Resolucion": "",
                    "Fecha_Resolucion": "",
                    "Link_Carpeta_Drive": fila["folder_link"],
                    "Links_Adjuntos": link_comprobante.strip(),
                },
                sheet_id=SHEET_ID,
            )
            clear_cache()
            _scan_tabla_1_completo.clear()
            st.session_state._t1_filas = None
            st.success("✅ Registro marcado como ENVIADO.")
            st.rerun()


@st.dialog("Actualizar caso")
def _dialog_actualizar_enviado(fila: dict):
    """
    BUG FIX: ahora extrae el file_id con múltiples estrategias y usa
    fallback a Drive API si el link no parsea. Nunca falla silenciosamente.
    """
    proveedor = str(fila.get("Proveedor", "")).strip()
    st.markdown(f"### Actualizar caso — {proveedor}")

    id_registro = str(fila.get("ID_Registro", "")).strip()
    numero_caso_default = str(fila.get("Numero_Caso_RMA", "")).strip()
    fecha_recepcion_raw = str(fila.get("Fecha_Recepcion", "")).strip()
    resolucion_default = str(fila.get("Resolucion", "")).strip()
    detalle_default = str(fila.get("Detalle_Resolucion", "")).strip()
    links_adjuntos_default = str(fila.get("Links_Adjuntos", "")).strip()

    fecha_recepcion_default = None
    if fecha_recepcion_raw:
        try:
            fecha_recepcion_default = datetime.strptime(fecha_recepcion_raw, "%Y-%m-%d").date()
        except Exception:
            pass

    numero_caso = st.text_input("Número de caso RMA", value=numero_caso_default, key=f"t2_nc_{id_registro}")
    fecha_recepcion = st.date_input(
        "Fecha de recepción", value=fecha_recepcion_default, format="DD/MM/YYYY", key=f"t2_fr_{id_registro}"
    )
    resolucion = st.selectbox(
        "Resolución",
        ["", "NOTA DE CREDITO", "CAMBIO", "RECHAZADO"],
        index=(["", "NOTA DE CREDITO", "CAMBIO", "RECHAZADO"].index(resolucion_default)
               if resolucion_default in ["", "NOTA DE CREDITO", "CAMBIO", "RECHAZADO"] else 0),
        key=f"t2_res_{id_registro}",
    )

    numero_nc, link_nc_pdf = "", ""
    if resolucion == "NOTA DE CREDITO":
        numero_nc = st.text_input("Número de Nota de Crédito", value=detalle_default, key=f"t2_nnc_{id_registro}").strip()
        link_nc_pdf = st.text_input("Link PDF de la Nota de Crédito (opcional)", value=links_adjuntos_default, key=f"t2_lnc_{id_registro}")

    if st.button("GUARDAR", type="primary", use_container_width=True, key=f"t2_save_{id_registro}"):
        if resolucion == "NOTA DE CREDITO" and not numero_nc:
            st.error("El número de Nota de Crédito es obligatorio.")
            return

        # ── Actualizar campos base en Sheets ──
        detalle_resolucion = numero_nc if resolucion == "NOTA DE CREDITO" else ""
        fecha_recep_str = fecha_recepcion.strftime("%Y-%m-%d") if fecha_recepcion else ""

        campos_sheets = {
            "Numero_Caso_RMA": numero_caso.strip(),
            "Fecha_Recepcion": fecha_recep_str,
            "Resolucion": resolucion,
            "Detalle_Resolucion": detalle_resolucion,
        }

        if resolucion == "NOTA DE CREDITO":
            nuevo_link_nc = link_nc_pdf.strip()
            actuales = [x.strip() for x in links_adjuntos_default.split(" | ") if x.strip()]
            if nuevo_link_nc and nuevo_link_nc not in actuales:
                actuales.append(nuevo_link_nc)
            campos_sheets["Links_Adjuntos"] = " | ".join(actuales)

        actualizar_registro(id_registro, campos_sheets, sheet_id=SHEET_ID)

        # ── Si hay resolución: mover archivo a /finalizado en Drive ──
        if resolucion:
            drive_client, _ = get_google_clients()

            # ── BUG FIX: múltiples estrategias para obtener file_id ──
            file_link_actual = str(fila.get("Link_Excel_Detalle", "")).strip()
            file_id = extraer_file_id_de_link(file_link_actual)

            # Fallback 1: usar file_id guardado en session si viene de Tabla 2
            if not file_id:
                file_id = str(fila.get("file_id", "")).strip()

            if not file_id:
                st.error(
                    f"❌ No se pudo extraer file_id del link: '{file_link_actual}'. "
                    "El registro fue guardado en la planilla pero el archivo NO fue movido en Drive. "
                    "Verificá el link en la columna Link_Excel_Detalle."
                )
                clear_cache()
                st.rerun()
                return

            # ── BUG FIX: obtener parent_id con fallback a Drive API ──
            carpeta_link = str(fila.get("Link_Carpeta_Drive", "")).strip()
            parent_id = extraer_folder_id_de_link(carpeta_link)

            if not parent_id:
                # Fallback: preguntar a Drive directamente por el parent actual
                parent_id = obtener_parent_desde_drive(drive_client, file_id)

            if not parent_id:
                st.error(
                    "❌ No se pudo obtener la carpeta del proveedor. "
                    "El registro fue guardado pero el archivo NO fue movido. "
                    "Verificá el link en la columna Link_Carpeta_Drive."
                )
                clear_cache()
                st.rerun()
                return

            # ── Renombrar y mover a /finalizado ──
            nombre_actual = str(fila.get("archivo_nombre", "")).strip()
            if not nombre_actual:
                nombre_actual = obtener_nombre_archivo(drive_client, file_id)

            if not nombre_actual:
                st.error("❌ No se pudo obtener el nombre del archivo en Drive.")
                return

            nuevo_nombre = _nombre_finalizado(nombre_actual, resolucion, detalle_resolucion)

            try:
                renombrar_archivo(drive_client, file_id, nuevo_nombre)
            except Exception as e:
                st.error(f"❌ Error al renombrar: {e}")
                return

            try:
                # El parent_id puede ser /procesado o la carpeta del proveedor directamente.
                # Necesitamos la carpeta del PROVEEDOR (parent del parent si ya está en /procesado).
                # Obtenemos el parent actual del archivo para determinar si hay que subir un nivel.
                parent_actual = obtener_parent_desde_drive(drive_client, file_id)

                # Asegurar carpeta /finalizado dentro del parent de nivel proveedor
                # Si el archivo está en /procesado, su parent es procesado, y el proveedor
                # es el parent de procesado. Aquí usamos parent_id que ya es la carpeta del proveedor.
                finalizado_id = asegurar_subcarpeta(drive_client, parent_id, "finalizado")
                mover_archivo(drive_client, file_id, finalizado_id)

                fupd = (
                    drive_client.files()
                    .get(fileId=file_id, fields="webViewLink", supportsAllDrives=True)
                    .execute()
                )
                nuevo_link = fupd.get("webViewLink", file_link_actual)
            except Exception as e:
                st.error(f"❌ Se renombró el archivo pero no se pudo mover a /finalizado: {e}")
                return

            actualizar_registro(
                id_registro,
                {
                    "Estado": "FINALIZADO",
                    "Fecha_Resolucion": date.today().strftime("%Y-%m-%d"),
                    "Link_Excel_Detalle": nuevo_link,
                },
                sheet_id=SHEET_ID,
            )

        clear_cache()
        _leer_excel_detalle_cache.clear()
        _get_file_meta_cached.clear()
        _scan_tabla_1_completo.clear()
        st.session_state._t1_filas = None
        st.success("✅ Registro actualizado correctamente.")
        st.rerun()


# ══════════════════════════════════════════════════════════
#  TABLA 1 — CARGADOS / POR ENVIAR
# ══════════════════════════════════════════════════════════

with st.expander("1. CARGADOS / POR ENVIAR", expanded=True):
    col_r1, col_ri, _ = st.columns([1, 1, 6])
    do_full_refresh = col_r1.button("🔄 Refresh completo", use_container_width=True, key="t1_refresh")
    do_incr_refresh = col_ri.button("⚡ Refresh rápido", use_container_width=True, key="t1_incr",
                                     help="Solo re-escanea archivos modificados desde el último refresh")

    if do_full_refresh:
        _scan_tabla_1_completo.clear()
        clear_cache()
        st.session_state._t1_filas = None
        st.rerun()

    # Escaneo completo (con cache @st.cache_data de 120s)
    if st.session_state._t1_filas is None:
        with st.spinner("Escaneando carpetas de proveedores…"):
            (
                st.session_state._t1_filas,
                st.session_state._t1_carpetas_validas,
                st.session_state._t1_carpetas_invalidas,
                st.session_state._t1_excels_invalidos,
                st.session_state._t1_last_scan_iso,
            ) = _scan_tabla_1_completo()

    # Refresh incremental
    if do_incr_refresh and st.session_state._t1_last_scan_iso:
        with st.spinner("Buscando archivos modificados…"):
            nuevas_filas, nuevos_invalidos = _scan_tabla_1_incremental(
                st.session_state._t1_carpetas_validas,
                st.session_state._t1_filas,
                st.session_state._t1_last_scan_iso,
            )
            st.session_state._t1_filas = nuevas_filas
            st.session_state._t1_excels_invalidos = (
                st.session_state._t1_excels_invalidos + nuevos_invalidos
            )
            st.session_state._t1_last_scan_iso = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S")
        st.rerun()

    filas = st.session_state._t1_filas or []
    carpetas_validas = st.session_state._t1_carpetas_validas
    carpetas_invalidas = st.session_state._t1_carpetas_invalidas
    excels_invalidos = st.session_state._t1_excels_invalidos

    st.caption(
        f"Carpetas válidas: **{len(carpetas_validas)}** | "
        f"Archivos pendientes: **{len(filas)}** | "
        f"Último scan: {st.session_state._t1_last_scan_iso or '—'} UTC"
    )

    if not filas:
        st.info("No hay archivos Excel cargados pendientes de enviar.")
    else:
        header = st.columns([2.2, 1.2, 1.1, 2.6, 1.0])
        for col, txt in zip(header, ["Proveedor", "Fecha", "Artículos", "Archivo", "Acción"]):
            col.markdown(f'<div class="tabla1-head-cell">{txt}</div>', unsafe_allow_html=True)
        st.markdown('<div class="tabla1-header-divider"></div>', unsafe_allow_html=True)

        for idx, fila in enumerate(filas):
            cols = st.columns([2.2, 1.2, 1.1, 2.6, 1.0])
            cols[0].write(fila["proveedor_mostrado"])
            cols[1].write(fila["fecha"])
            cant = str(fila["cantidad_articulos"]) if fila["cantidad_articulos"] != "" else "0"
            if cols[2].button(cant, key=f"t1_det_{idx}", use_container_width=True):
                _dialog_detalle_articulos(fila["archivo_nombre"], fila["archivo_id"])
            cols[3].markdown(f'[{fila["archivo_nombre"]}]({fila["archivo_link"]})')
            if cols[4].button("ENVIADO", key=f"t1_env_{idx}", use_container_width=True):
                _dialog_enviado(fila)
            st.markdown('<div class="tabla1-row-divider"></div>', unsafe_allow_html=True)

    with st.expander("Archivos/carpetas ignorados", expanded=False):
        if carpetas_invalidas:
            st.markdown("**Carpetas ignoradas**")
            st.dataframe(pd.DataFrame(carpetas_invalidas), use_container_width=True, hide_index=True)
        if excels_invalidos:
            st.markdown("**Excels ignorados**")
            st.dataframe(pd.DataFrame(excels_invalidos), use_container_width=True, hide_index=True)
        if not carpetas_invalidas and not excels_invalidos:
            st.caption("Sin elementos ignorados.")


# ══════════════════════════════════════════════════════════
#  TABLA 2 — ENVIADOS / EN PROCESO
# ══════════════════════════════════════════════════════════

with st.expander("2. ENVIADOS / EN PROCESO", expanded=True):
    c2, _ = st.columns([1, 7])
    if c2.button("🔄 Refresh", key="t2_refresh", use_container_width=True):
        clear_cache()
        _leer_excel_detalle_cache.clear()
        _get_file_meta_cached.clear()
        st.rerun()

    df_enviados = leer_registros(estado="ENVIADO", sheet_id=SHEET_ID)

    if df_enviados.empty:
        st.info("No hay registros ENVIADOS.")
    else:
        # Enriquecer con nombre de archivo (cacheado por file_id)
        filas_t2 = []
        for _, r in df_enviados.iterrows():
            fila = r.to_dict()
            file_id = extraer_file_id_de_link(str(fila.get("Link_Excel_Detalle", "")))
            meta = _get_file_meta_cached(file_id) if file_id else {}
            fila["archivo_nombre"] = meta.get("name", "")
            fila["file_id"] = file_id
            filas_t2.append(fila)

        headers_t2 = ["Proveedor", "Fecha envío", "Artículos", "N° envío",
                       "N° caso RMA", "Fecha recepción", "Resolución", "Archivo", "Acción"]
        widths_t2 = [2.2, 1.1, 0.9, 1.2, 1.4, 1.2, 1.4, 1.3, 1.1]
        hcols = st.columns(widths_t2)
        for hc, ht in zip(hcols, headers_t2):
            hc.markdown(f'<div class="tabla1-head-cell">{ht}</div>', unsafe_allow_html=True)
        st.markdown('<div class="tabla1-header-divider"></div>', unsafe_allow_html=True)

        for idx, fila in enumerate(filas_t2):
            cols = st.columns(widths_t2)
            cols[0].write(str(fila.get("Proveedor", "")))
            cols[1].write(_to_ddmmyyyy_safe(str(fila.get("Fecha_Envio", ""))))
            cant2 = str(fila.get("Cantidad_Articulos", "") or "0")
            file_id_t2 = str(fila.get("file_id", "")).strip()
            if cols[2].button(cant2, key=f"t2_det_{idx}", use_container_width=True):
                if file_id_t2:
                    _dialog_detalle_articulos(fila.get("archivo_nombre", "Detalle"), file_id_t2)
                else:
                    st.toast("No se pudo obtener el archivo.", icon="⚠️")
            cols[3].write(str(fila.get("Numero_Envio", "")))
            cols[4].write(str(fila.get("Numero_Caso_RMA", "")))
            cols[5].write(_to_ddmmyyyy_safe(str(fila.get("Fecha_Recepcion", ""))))
            cols[6].write(str(fila.get("Resolucion", "")))
            link_t2 = str(fila.get("Link_Excel_Detalle", ""))
            if link_t2:
                cols[7].markdown(f"[Archivo]({link_t2})")
            else:
                cols[7].write("—")
            if cols[8].button("ACTUALIZAR", key=f"t2_upd_{fila.get('ID_Registro', idx)}", use_container_width=True):
                _dialog_actualizar_enviado(fila)
            st.markdown('<div class="tabla1-row-divider"></div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
#  TABLA 3 — FINALIZADOS (HISTÓRICO)
# ══════════════════════════════════════════════════════════

with st.expander("3. FINALIZADOS (HISTÓRICO)", expanded=True):
    c3, _ = st.columns([1, 7])
    if c3.button("🔄 Refresh", key="t3_refresh", use_container_width=True):
        clear_cache()
        st.rerun()

    df_fin = leer_registros(estado="FINALIZADO", sheet_id=SHEET_ID)

    if not df_fin.empty:
        # ── Filtros ──
        st.markdown("**Filtros**")
        fc1, fc2, fc3, fc4 = st.columns([2, 1.5, 1.5, 1.5])
        filtro_prov = fc1.text_input("Proveedor", placeholder="Buscar…", key="t3_fp",
                                      label_visibility="collapsed")
        filtro_res = fc2.selectbox("Resolución", ["Todas", "NOTA DE CREDITO", "CAMBIO", "RECHAZADO"],
                                    key="t3_fr", label_visibility="collapsed")
        filtro_desde = fc3.date_input("Desde resolución", value=None, format="DD/MM/YYYY",
                                       key="t3_fd", label_visibility="collapsed")
        filtro_hasta = fc4.date_input("Hasta resolución", value=None, format="DD/MM/YYYY",
                                       key="t3_fh", label_visibility="collapsed")

        df_show = df_fin.copy()
        if filtro_prov.strip():
            df_show = df_show[df_show["Proveedor"].str.contains(filtro_prov.strip(), case=False, na=False)]
        if filtro_res != "Todas":
            df_show = df_show[df_show["Resolucion"].str.strip() == filtro_res]
        if filtro_desde:
            df_show = df_show[df_show["Fecha_Resolucion"].str.strip() >= filtro_desde.strftime("%Y-%m-%d")]
        if filtro_hasta:
            df_show = df_show[df_show["Fecha_Resolucion"].str.strip() <= filtro_hasta.strftime("%Y-%m-%d")]

        st.caption(f"Total histórico: **{len(df_fin)}** | Mostrando: **{len(df_show)}**")

        if df_show.empty:
            st.info("No hay registros que coincidan con los filtros.")
        else:
            headers_t3 = [
                "Proveedor", "Fecha carga", "Fecha envío", "N° envío",
                "N° caso RMA", "Resolución", "Detalle", "Fecha resolución", "Archivo"
            ]
            widths_t3 = [2.0, 1.0, 1.0, 1.1, 1.3, 1.4, 1.2, 1.2, 1.0]
            hcols3 = st.columns(widths_t3)
            for hc, ht in zip(hcols3, headers_t3):
                hc.markdown(f'<div class="tabla1-head-cell">{ht}</div>', unsafe_allow_html=True)
            st.markdown('<div class="tabla1-header-divider"></div>', unsafe_allow_html=True)

            for idx, (_, row) in enumerate(df_show.iterrows()):
                cols3 = st.columns(widths_t3)
                cols3[0].write(str(row.get("Proveedor", "")))
                cols3[1].write(_to_ddmmyyyy_safe(str(row.get("Fecha_carga", ""))))
                cols3[2].write(_to_ddmmyyyy_safe(str(row.get("Fecha_Envio", ""))))
                cols3[3].write(str(row.get("Numero_Envio", "")))
                cols3[4].write(str(row.get("Numero_Caso_RMA", "")))
                res_val = str(row.get("Resolucion", ""))
                cols3[5].markdown(_badge_resolucion(res_val), unsafe_allow_html=True)
                cols3[6].write(str(row.get("Detalle_Resolucion", "")))
                cols3[7].write(_to_ddmmyyyy_safe(str(row.get("Fecha_Resolucion", ""))))
                link_t3 = str(row.get("Link_Excel_Detalle", ""))
                if link_t3:
                    cols3[8].markdown(f"[Archivo]({link_t3})")
                else:
                    cols3[8].write("—")
                st.markdown('<div class="tabla1-row-divider"></div>', unsafe_allow_html=True)

        # ── Exportar a Excel ──
        st.markdown("---")
        if not df_show.empty and st.button("📥 Exportar a Excel", key="t3_export"):
            export_cols = [
                "Proveedor", "Fecha_carga", "Fecha_Envio", "Numero_Envio",
                "Numero_Caso_RMA", "Resolucion", "Detalle_Resolucion",
                "Fecha_Resolucion", "Links_Adjuntos",
            ]
            df_exp = df_show[[c for c in export_cols if c in df_show.columns]].copy()
            for c in ("Fecha_carga", "Fecha_Envio", "Fecha_Resolucion"):
                if c in df_exp.columns:
                    df_exp[c] = df_exp[c].apply(_to_ddmmyyyy_safe)
            buf_exp = io.BytesIO()
            with pd.ExcelWriter(buf_exp, engine="xlsxwriter") as writer:
                df_exp.to_excel(writer, index=False, sheet_name="Finalizados")
            st.download_button(
                label="⬇️ Descargar archivo",
                data=buf_exp.getvalue(),
                file_name=f"RMA_Finalizados_{date.today().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="t3_dl",
            )
    else:
        st.info("No hay casos finalizados aún.")


# ══════════════════════════════════════════════════════════
#  HELPERS TEMPLATE DRIVE (Tabla 4 — Garantía)
# ══════════════════════════════════════════════════════════

@st.cache_data(ttl=300, show_spinner=False)
def _buscar_template_en_carpeta_cache(folder_id: str) -> dict:
    """Busca template.xlsx y template.txt en la carpeta del proveedor en Drive."""
    try:
        drive_client, _ = get_google_clients()
        response = drive_client.files().list(
            q=(
                f"'{folder_id}' in parents and trashed=false "
                f"and (name='template.xlsx' or name='template.txt')"
            ),
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        files = response.get("files", [])
        return {
            "xlsx_id": next((f["id"] for f in files if f["name"] == "template.xlsx"), None),
            "txt_id":  next((f["id"] for f in files if f["name"] == "template.txt"),  None),
        }
    except Exception:
        return {"xlsx_id": None, "txt_id": None}


def _buscar_carpeta_proveedor(proveedor_nombre: str) -> str | None:
    """
    Devuelve el folder_id de la carpeta del proveedor en Drive.
    Busca primero en session_state (_t1_carpetas_validas), luego directo en Drive.
    """
    carpetas = st.session_state.get("_t1_carpetas_validas", [])
    nombre_p = proveedor_nombre.lower().strip()
    for c in carpetas:
        nombre_c = c.get("nombre", "").lower().strip()
        if nombre_p in nombre_c or nombre_c in nombre_p:
            return c["folder_id"]
    # Fallback: consultar Drive directamente
    try:
        drive_client, _ = get_google_clients()
        resp = drive_client.files().list(
            q=(
                f"'{DRIVE_ROOT_FOLDER_ID}' in parents "
                f"and mimeType='application/vnd.google-apps.folder' "
                f"and trashed=false"
            ),
            fields="files(id, name)",
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
            pageSize=200,
        ).execute()
        for f in resp.get("files", []):
            nombre_f = f["name"].lower().strip()
            if nombre_p in nombre_f or nombre_f in nombre_p:
                return f["id"]
    except Exception:
        pass
    return None


def _descargar_bytes_drive(file_id: str) -> bytes | None:
    """Descarga un archivo de Drive y lo devuelve como bytes."""
    try:
        from googleapiclient.http import MediaIoBaseDownload
        drive_client, _ = get_google_clients()
        request = drive_client.files().get_media(fileId=file_id, supportsAllDrives=True)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return fh.getvalue()
    except Exception:
        return None


def _parsear_template_txt(contenido: str) -> dict | None:
    """
    Parsea template.txt con el formato:
        Producto         = A24
        Serial           = B24
        diagnostico      = D24
        ProductoFinal    = A41
        SerialFinal      = B41
        diagnosticoFinal = D41

    Retorna dict con:
        campos:       {"Producto": ("A", 24), "Serial": ("B", 24), ...}
        primera_fila: int  (ej. 24)
        ultima_fila:  int  (ej. 41)
        capacidad:    int  (ej. 18 — artículos por planilla)
    """
    config = {}
    for line in contenido.strip().splitlines():
        line = line.strip()
        if not line or "=" not in line:
            continue
        key, _, val = line.partition("=")
        key = key.strip()
        val = val.strip().upper()
        m = re.match(r"([A-Z]+)(\d+)", val)
        if m:
            config[key] = (m.group(1), int(m.group(2)))

    if not config:
        return None

    campos_base  = {k: v for k, v in config.items() if not k.lower().endswith("final")}
    campos_final = {k: v for k, v in config.items() if     k.lower().endswith("final")}

    if not campos_base:
        return None

    primera_fila = min(v[1] for v in campos_base.values())
    ultima_fila  = max(v[1] for v in campos_final.values()) if campos_final else primera_fila

    return {
        "campos":       campos_base,
        "primera_fila": primera_fila,
        "ultima_fila":  ultima_fila,
        "capacidad":    ultima_fila - primera_fila + 1,
    }


def _generar_excels_con_template(
    template_bytes: bytes,
    template_config: dict,
    registros: list,
    proveedor: str,
) -> list[tuple[str, bytes]]:
    """
    Rellena el template con los registros de Airtable.
    Si los registros superan la capacidad del template, genera múltiples archivos.
    Retorna lista de (nombre_archivo, bytes).
    """
    from openpyxl import load_workbook

    capacidad    = template_config["capacidad"]
    primera_fila = template_config["primera_fila"]
    campos       = template_config["campos"]  # {"Producto": ("A", 24), ...}

    chunks = [registros[i:i + capacidad] for i in range(0, len(registros), capacidad)]
    resultados = []

    for parte_idx, chunk in enumerate(chunks, 1):
        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active

        for item_idx, registro in enumerate(chunk):
            fila = primera_fila + item_idx
            for campo_nombre, (col_letter, _) in campos.items():
                valor = registro.get(campo_nombre, "")
                ws[f"{col_letter}{fila}"] = valor if valor is not None else ""

        buf = io.BytesIO()
        wb.save(buf)

        sufijo = f"_parte{parte_idx:02d}" if len(chunks) > 1 else ""
        nombre = (
            f"Garantia_{proveedor.replace(' ', '_')}"
            f"_{date.today().strftime('%Y%m%d')}{sufijo}.xlsx"
        )
        resultados.append((nombre, buf.getvalue()))

    return resultados


# ══════════════════════════════════════════════════════════
#  HERRAMIENTA: VERIFICAR PRODUCTOS PARA GARANTÍA (Airtable)
# ══════════════════════════════════════════════════════════

with st.expander("🔍 VERIFICAR PRODUCTOS PARA GARANTÍA", expanded=False):

    st.markdown(
        "Consultá Airtable para encontrar productos pendientes de garantía de un proveedor "
        "y generá el Excel de remisión. Al confirmar, los registros quedan marcados como capturados."
    )

    # ── API Key de Airtable (guardada en secrets) ──────────────────────────────
    try:
        AIRTABLE_API_KEY = st.secrets["AIRTABLE"]["API_KEY"]
    except Exception:
        st.error(
            "❌ No se encontró la clave de Airtable en Secrets. "
            "Agregá `[AIRTABLE]` con `API_KEY = '...'` en `.streamlit/secrets.toml`."
        )
        AIRTABLE_API_KEY = None

    if AIRTABLE_API_KEY:

        col_at1, _ = st.columns([2, 6])

        @st.cache_data(ttl=120, show_spinner=False)
        def _listar_proveedores_airtable() -> list[str]:
            return listar_proveedores_unicos(AIRTABLE_API_KEY)

        with col_at1:
            if st.button("🔄 Actualizar lista de proveedores", use_container_width=True, key="at_ref_prov"):
                _listar_proveedores_airtable.clear()
                st.rerun()

        try:
            lista_proveedores_at = _listar_proveedores_airtable()
        except Exception as exc:
            st.error(f"❌ Error al conectar con Airtable: {exc}")
            lista_proveedores_at = []

        if not lista_proveedores_at:
            st.warning("No se encontraron proveedores en Airtable o la conexión falló.")
        else:
            proveedor_sel = st.selectbox(
                "Seleccioná un proveedor:",
                options=["— Elegir —"] + lista_proveedores_at,
                key="at_prov_sel",
            )

            if proveedor_sel and proveedor_sel != "— Elegir —":

                if st.button(
                    f"🔎 Buscar productos pendientes de «{proveedor_sel}»",
                    key="at_buscar",
                ):
                    with st.spinner("Consultando Airtable…"):
                        try:
                            registros_at = obtener_registros_proveedor(AIRTABLE_API_KEY, proveedor_sel)
                            st.session_state["_at_registros"] = registros_at
                            st.session_state["_at_proveedor"] = proveedor_sel
                        except Exception as exc:
                            st.error(f"❌ Error al consultar Airtable: {exc}")
                            st.session_state.pop("_at_registros", None)

                if (
                    "_at_registros" in st.session_state
                    and st.session_state.get("_at_proveedor") == proveedor_sel
                ):
                    registros_at = st.session_state["_at_registros"]

                    if not registros_at:
                        st.info(
                            f"No hay productos pendientes de captura para «{proveedor_sel}»."
                        )
                    else:
                        st.success(
                            f"Se encontraron **{len(registros_at)}** producto(s) pendiente(s) "
                            f"para «{proveedor_sel}»."
                        )

                        df_at = pd.DataFrame(registros_at)[["Producto", "Serial", "diagnostico"]]
                        st.dataframe(df_at, use_container_width=True, hide_index=True)

                        def _generar_excel_garantia(registros: list, proveedor: str) -> bytes:
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            from openpyxl.utils import get_column_letter

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Garantia"

                            COLOR_HDR_BG = "1F4E79"
                            COLOR_HDR_FG = "FFFFFF"
                            COLOR_ALT    = "D6E4F0"

                            thin = Side(style="thin", color="AAAAAA")
                            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

                            # Título
                            ws.merge_cells("A1:C1")
                            c = ws["A1"]
                            c.value = f"Verificación de Garantía — {proveedor}"
                            c.font  = Font(name="Arial", bold=True, size=13, color=COLOR_HDR_FG)
                            c.fill  = PatternFill("solid", start_color=COLOR_HDR_BG)
                            c.alignment = Alignment(horizontal="center", vertical="center")
                            ws.row_dimensions[1].height = 26

                            # Fecha
                            ws.merge_cells("A2:C2")
                            c2 = ws["A2"]
                            c2.value = f"Generado: {date.today().strftime('%d/%m/%Y')}"
                            c2.font  = Font(name="Arial", size=9, italic=True, color="555555")
                            c2.alignment = Alignment(horizontal="center")

                            # Encabezados
                            for ci, enc in enumerate(["Producto", "Serial", "Diagnóstico"], 1):
                                c = ws.cell(row=4, column=ci, value=enc)
                                c.font  = Font(name="Arial", bold=True, size=11, color=COLOR_HDR_FG)
                                c.fill  = PatternFill("solid", start_color=COLOR_HDR_BG)
                                c.alignment = Alignment(horizontal="center", vertical="center")
                                c.border = brd
                            ws.row_dimensions[4].height = 20

                            # Datos
                            for i, rec in enumerate(registros, start=5):
                                fill = PatternFill("solid", start_color=COLOR_ALT) if i % 2 == 0 else None
                                for ci, campo in enumerate(["Producto", "Serial", "diagnostico"], 1):
                                    c = ws.cell(row=i, column=ci, value=rec.get(campo, ""))
                                    c.font = Font(name="Arial", size=10)
                                    c.alignment = Alignment(vertical="center", wrap_text=True)
                                    c.border = brd
                                    if fill:
                                        c.fill = fill
                                ws.row_dimensions[i].height = 16

                            # Total
                            ft = len(registros) + 5
                            ws.merge_cells(f"A{ft}:B{ft}")
                            ws.cell(row=ft, column=1, value="Total productos").font = Font(name="Arial", bold=True, size=10)
                            ws.cell(row=ft, column=1).alignment = Alignment(horizontal="right")
                            ws.cell(row=ft, column=3, value=len(registros)).font = Font(name="Arial", bold=True, size=10)
                            ws.cell(row=ft, column=3).alignment = Alignment(horizontal="center")

                            for ci, ancho in enumerate([45, 22, 50], 1):
                                ws.column_dimensions[get_column_letter(ci)].width = ancho

                            buf = io.BytesIO()
                            wb.save(buf)
                            return buf.getvalue()

                        # ── Buscar template del proveedor en Drive ────────────
                        _folder_id_prov = _buscar_carpeta_proveedor(proveedor_sel)
                        _tmpl_info = (
                            _buscar_template_en_carpeta_cache(_folder_id_prov)
                            if _folder_id_prov
                            else {"xlsx_id": None, "txt_id": None}
                        )
                        _usa_template = bool(
                            _tmpl_info.get("xlsx_id") and _tmpl_info.get("txt_id")
                        )
                        archivos_excel_descarga = []

                        if _usa_template:
                            with st.spinner("Cargando template desde Drive…"):
                                _tmpl_xlsx_b = _descargar_bytes_drive(_tmpl_info["xlsx_id"])
                                _tmpl_txt_b  = _descargar_bytes_drive(_tmpl_info["txt_id"])
                                _tmpl_cfg = (
                                    _parsear_template_txt(
                                        _tmpl_txt_b.decode("utf-8", errors="replace")
                                    )
                                    if _tmpl_txt_b else None
                                )
                            if _tmpl_xlsx_b and _tmpl_cfg:
                                archivos_excel_descarga = _generar_excels_con_template(
                                    _tmpl_xlsx_b, _tmpl_cfg, registros_at, proveedor_sel
                                )
                                n_arch = len(archivos_excel_descarga)
                                cap    = _tmpl_cfg["capacidad"]
                                st.info(
                                    f"📋 Template del proveedor encontrado — "
                                    f"capacidad: **{cap}** artículos/planilla."
                                    + (f" Se generarán **{n_arch}** archivos." if n_arch > 1 else "")
                                )
                            else:
                                st.warning(
                                    "⚠️ Template encontrado en Drive pero no pudo procesarse. "
                                    "Se usará el formato genérico."
                                )

                        if not archivos_excel_descarga:
                            # Fallback: Excel genérico
                            archivos_excel_descarga = [
                                (
                                    f"Garantia_{proveedor_sel.replace(' ', '_')}"
                                    f"_{date.today().strftime('%Y%m%d')}.xlsx",
                                    _generar_excel_garantia(registros_at, proveedor_sel),
                                )
                            ]

                        # ── Descarga y confirmación ───────────────────────────
                        st.markdown("---")
                        st.markdown("**¿Qué querés hacer con estos registros?**")
                        col_dl, col_conf, _ = st.columns([1.5, 2, 4.5])

                        if len(archivos_excel_descarga) == 1:
                            _nombre_dl, _bytes_dl = archivos_excel_descarga[0]
                            col_dl.download_button(
                                label="⬇️ Descargar Excel",
                                data=_bytes_dl,
                                file_name=_nombre_dl,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="at_dl",
                                use_container_width=True,
                            )
                        else:
                            import zipfile as _zipfile
                            _zip_buf = io.BytesIO()
                            with _zipfile.ZipFile(_zip_buf, "w", _zipfile.ZIP_DEFLATED) as _zf:
                                for _nf, _bf in archivos_excel_descarga:
                                    _zf.writestr(_nf, _bf)
                            _zip_name = (
                                f"Garantia_{proveedor_sel.replace(' ', '_')}"
                                f"_{date.today().strftime('%Y%m%d')}.zip"
                            )
                            col_dl.download_button(
                                label=f"⬇️ Descargar ZIP ({len(archivos_excel_descarga)} archivos)",
                                data=_zip_buf.getvalue(),
                                file_name=_zip_name,
                                mime="application/zip",
                                key="at_dl_zip",
                                use_container_width=True,
                            )
                            with st.expander(
                                f"📂 Descargar archivos individuales "
                                f"({len(archivos_excel_descarga)})",
                                expanded=False,
                            ):
                                for _i, (_nf, _bf) in enumerate(archivos_excel_descarga):
                                    st.download_button(
                                        label=f"⬇️ {_nf}",
                                        data=_bf,
                                        file_name=_nf,
                                        mime=(
                                            "application/vnd.openxmlformats-officedocument"
                                            ".spreadsheetml.sheet"
                                        ),
                                        key=f"at_dl_ind_{_i}",
                                        use_container_width=True,
                                    )

                        if col_conf.button(
                            "✅ Confirmar y marcar como capturados",
                            key="at_confirmar",
                            use_container_width=True,
                            type="primary",
                        ):
                            ids = [r["id"] for r in registros_at]
                            with st.spinner(f"Marcando {len(ids)} registros en Airtable…"):
                                try:
                                    cant, errores = marcar_como_capturados(AIRTABLE_API_KEY, ids)
                                except Exception as exc:
                                    st.error(f"❌ Error al actualizar Airtable: {exc}")
                                    cant, errores = 0, [str(exc)]

                            if errores:
                                st.warning(
                                    f"Se actualizaron {cant} registro(s) pero hubo "
                                    f"{len(errores)} error(es): {'; '.join(errores[:3])}"
                                )
                            else:
                                st.success(
                                    f"✅ {cant} registro(s) marcados como capturados en Airtable. "
                                    "No aparecerán en futuras consultas."
                                )
                            st.session_state.pop("_at_registros", None)
                            st.session_state.pop("_at_proveedor", None)
                            _listar_proveedores_airtable.clear()
                            st.rerun()
